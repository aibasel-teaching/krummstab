import logging
from collections import defaultdict
import itertools
from pathlib import Path

from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula


from .. import config, sheets, strings, utils
from ..teams import *


def convert_to_float_if_possible(value):
    try:
        return float(value)
    except ValueError:
        return value


def total_score(values):
    total = 0
    for v in values:
        # Marks are either per sheet (one value per sheet)
        # or per exercise (one dictionary per sheet).
        if isinstance(v, dict):
            total += total_score(v.values())
        else:
            try:
                total += float(v)
            except ValueError:
                pass
    return total


def sort_items_by_score(email_scores_pair):
    scores = email_scores_pair[1].values()
    return total_score(scores)


def first_not_none(*values):
    """
    Returns the first value different from None.
    """
    for v in values:
        if v is not None:
            return v


def merge_borders(b1: Border | None, b2: Border | None):
    if not b1:
        return b2
    if not b2:
        return b1
    return Border(
        left=first_not_none(b2.left, b1.left),
        right=first_not_none(b2.right or b1.right),
        top=first_not_none(b2.top, b1.top),
        bottom=first_not_none(b2.bottom, b1.bottom),
        diagonal=first_not_none(b2.diagonal, b1.diagonal),
        diagonal_direction=first_not_none(
            b2.diagonal_direction, b1.diagonal_direction
        ),
        outline=first_not_none(b2.outline, b1.outline),
        vertical=first_not_none(b2.vertical, b1.vertical),
        horizontal=first_not_none(b2.horizontal, b1.horizontal),
    )


def merge_alignment(a1: Alignment | None, a2: Alignment | None):
    if not a1:
        return a2
    if not a2:
        return a1
    return Alignment(
        horizontal=first_not_none(a2.horizontal, a1.horizontal),
        vertical=first_not_none(a2.vertical, a1.vertical),
        wrap_text=first_not_none(a2.wrap_text, a1.wrap_text),
        shrink_to_fit=first_not_none(a2.shrink_to_fit, a1.shrink_to_fit),
        text_rotation=first_not_none(a2.text_rotation, a1.text_rotation),
    )


class OpenpyxlRangeRef:
    def __init__(
        self,
        min_row,
        min_column,
        max_row=None,
        max_column=None,
        row_absolute=True,
        column_absolute=True,
    ):
        self.min_row = min_row
        self.min_column = min_column
        self.max_row = max_row or min_row
        self.max_column = max_column or min_column
        self.row_absolute = row_absolute
        self.column_absolute = column_absolute

    def __str__(self):
        def cell_ref(row, row_absolute, column, column_absolute):
            row_prefix = "$" if row_absolute else ""
            column_prefix = "$" if column_absolute else ""
            column_letter = get_column_letter(column)
            return f"{column_prefix}{column_letter}{row_prefix}{row}"

        start_cell = cell_ref(
            self.min_row,
            self.row_absolute,
            self.min_column,
            self.column_absolute,
        )
        if self.max_row == self.min_row and self.max_column == self.min_column:
            return start_cell
        else:
            end_cell = cell_ref(
                self.max_row,
                self.row_absolute,
                self.max_column,
                self.column_absolute,
            )
            return f"{start_cell}:{end_cell}"


class OpenpyxlStyle:
    def __init__(
        self,
        font=None,
        fill=None,
        border=None,
        alignment=None,
        number_format=None,
    ):
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.number_format = number_format

    def __or__(self, other):
        # We could introduce a merge_font, merge_fill, and merge_number_format
        # methods but we don't need them at the moment.
        return OpenpyxlStyle(
            font=other.font or self.font,
            fill=other.fill or self.fill,
            border=merge_borders(self.border, other.border),
            alignment=merge_alignment(self.alignment, other.alignment),
            number_format=other.number_format or self.number_format,
        )

    def apply_to(self, cell):
        if self.font:
            cell.font = self.font
        if self.fill:
            cell.fill = self.fill
        if self.border:
            cell.border = self.border
        if self.alignment:
            cell.alignment = self.alignment
        if self.number_format:
            cell.number_format = self.number_format

    def as_differential_style(self):
        return DifferentialStyle(
            font=self.font,
            fill=self.fill,
            border=self.border,
            alignment=self.alignment,
            numFmt=self.number_format,
        )


# Fonts
BOLD = OpenpyxlStyle(font=Font(bold=True))

# Borders
THIN = Side(style="thin")
BORDER_LEFT = OpenpyxlStyle(border=Border(left=THIN))
BORDER_RIGHT = OpenpyxlStyle(border=Border(right=THIN))
BORDER_TOP = OpenpyxlStyle(border=Border(top=THIN))
BORDER_BOTTOM = OpenpyxlStyle(border=Border(bottom=THIN))
BORDER_ALL = BORDER_LEFT | BORDER_RIGHT | BORDER_TOP | BORDER_BOTTOM

# Fills
GRAY = OpenpyxlStyle(fill=PatternFill(start_color="E2E2E2", fill_type="solid"))
GREEN = OpenpyxlStyle(fill=PatternFill(start_color="9BE189", fill_type="solid"))
YELLOW = OpenpyxlStyle(
    fill=PatternFill(start_color="FFE699", fill_type="solid")
)
RED = OpenpyxlStyle(fill=PatternFill(start_color="EE7868", fill_type="solid"))
PLAGIARISM_RED = OpenpyxlStyle(
    fill=PatternFill(start_color="FF0000", fill_type="solid")
)

# Number formats
PERCENT = OpenpyxlStyle(number_format="0%")
TWO_DIGIT_FLOAT = OpenpyxlStyle(number_format="0.00")

# Alignment
TEXT_WRAP = OpenpyxlStyle(alignment=Alignment(wrap_text=True))
CENTERED = OpenpyxlStyle(alignment=Alignment(horizontal="center"))
RIGHT_ALIGNED = OpenpyxlStyle(alignment=Alignment(horizontal="right"))

# Common styles
HEADING = BOLD | GRAY | CENTERED


class PointsSummarySheetBuilder:
    SHEET_NAME_SUMMARY = "Summary"
    SHEET_NAME_INDIVIDUAL = "Points per exercise"

    VAR_MIN_GOOD_IMPROVEMENT = "min_good_improvement"
    VAR_MIN_BAD_IMPROVEMENT = "min_bad_improvement"
    VAR_MARKS_BY_SHEET = "marks_by_sheet"
    VAR_SHEET_WAS_MARKED = "was_marked"
    VAR_MARKS_ALL_SHEETS = "marks_available_all_sheet"
    VAR_MARKS_PAST_SHEETS = "marks_available_past_sheet"
    VAR_MARKS_FUTURE_SHEETS = "marks_available_future_sheet"
    VAR_MARKS_TO_PASS = "marks_to_pass"

    def __init__(
        self,
        workbook,
        points_per,
        max_points_per_sheet,
        email_to_name,
        students_marks,
        graded_sheet_names,
    ):
        # Fix an order of the sheets and make sure available_marks is consistent
        # with it.
        self.points_per = points_per
        self.sheets = max_points_per_sheet.keys()
        self.available_marks = [
            max_points_per_sheet.get(sheet) for sheet in self.sheets
        ]
        self.was_marked = [
            int(sheet in graded_sheet_names) for sheet in self.sheets
        ]
        self.email_to_name = email_to_name
        self.students_marks = students_marks
        self.graded_sheet_names = graded_sheet_names
        self.workbook = workbook
        self.worksheet = None

    def write(self, row, col, value, format=None):
        cell = self.worksheet.cell(row, col, value)
        cell.value = value
        if format:
            format.apply_to(cell)
        return OpenpyxlRangeRef(row, col)

    def write_formula(self, row, col, formula, format=None):
        return self.write(row, col, formula, format)

    def write_array_formula(self, row, col, formula, format=None):
        cellref_target = OpenpyxlRangeRef(
            row, col, row_absolute=False, column_absolute=False
        )
        return self.write(
            row, col, ArrayFormula(str(cellref_target), formula), format
        )

    def merge_range(
        self, start_row, start_column, end_row, end_column, value, format=None
    ):
        ref = self.write(start_row, start_column, value, format)
        self.worksheet.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )
        return ref

    def write_row(self, row, col, values, format=None):
        for current_col, value in enumerate(values, start=col):
            self.write(row, current_col, value, format)
        return OpenpyxlRangeRef(row, col, row, col + len(values) - 1)

    def define_name(self, var, range: OpenpyxlRangeRef):
        range.row_absolute = True
        range.column_absolute = True
        defined_name = DefinedName(
            var, attr_text=f"{quote_sheetname(self.worksheet.title)}!{range}"
        )
        self.worksheet.defined_names.add(defined_name)

    def add_conditional_format(self, range, formula, format: OpenpyxlStyle):
        rule = Rule(
            type="expression",
            dxf=format.as_differential_style(),
            formula=[formula],
        )
        self.worksheet.conditional_formatting.add(str(range), rule)

    def write_sheet_name_row(self, row, col):
        self.write_row(row, col, self.sheets, BOLD | BORDER_TOP | BORDER_BOTTOM)

    def write_was_marked_row(self, row, col):
        self.merge_range(
            row, col, row, col + 1, "Was marked", BOLD | RIGHT_ALIGNED
        )
        ref = self.write_row(row, col + 2, self.was_marked)
        self.define_name(self.VAR_SHEET_WAS_MARKED, ref)

    def write_available_marks_by_sheet(self, row, col):
        self.merge_range(
            row, col, row, col + 1, "Available marks", BOLD | RIGHT_ALIGNED
        )
        ref = self.write_row(row, col + 2, self.available_marks)
        self.define_name(self.VAR_MARKS_BY_SHEET, ref)

    def write_available_marks_all_sheets(self, row, col):
        self.write(row, col, "All Sheets", HEADING | BORDER_TOP | BORDER_BOTTOM)
        ref = self.write_formula(
            row + 1,
            col,
            f"=SUM({self.VAR_MARKS_BY_SHEET})",
            HEADING | BORDER_BOTTOM,
        )
        self.define_name(self.VAR_MARKS_ALL_SHEETS, ref)

    def write_available_marks_past_sheets(self, row, col):
        self.write(
            row, col, "Past Sheets", HEADING | BORDER_TOP | BORDER_BOTTOM
        )
        ref = self.write_formula(
            row + 1,
            col,
            f"=SUMPRODUCT({self.VAR_MARKS_BY_SHEET},{self.VAR_SHEET_WAS_MARKED})",
            HEADING | BORDER_BOTTOM,
        )
        self.define_name(self.VAR_MARKS_PAST_SHEETS, ref)

    def write_available_marks_future_sheets(self, row, col):
        self.write(
            row,
            col,
            "Future Sheets",
            HEADING | BORDER_TOP | BORDER_BOTTOM | BORDER_RIGHT,
        )
        ref = self.write_formula(
            row + 1,
            col,
            f"={self.VAR_MARKS_ALL_SHEETS}-{self.VAR_MARKS_PAST_SHEETS}",
            HEADING | BORDER_BOTTOM | BORDER_RIGHT,
        )
        self.define_name(self.VAR_MARKS_FUTURE_SHEETS, ref)

    def write_available_marks_table(self, row, col):
        self.merge_range(
            row, col, row, col + 2, "Available Marks", HEADING | BORDER_ALL
        )
        self.write_available_marks_all_sheets(row + 1, col)
        self.write_available_marks_past_sheets(row + 1, col + 1)
        self.write_available_marks_future_sheets(row + 1, col + 2)

    def write_required_marks_table(self, row, col):
        self.write(
            row,
            col,
            "Marks required",
            HEADING | BORDER_LEFT | BORDER_RIGHT | TEXT_WRAP,
        )
        self.write(
            row + 1,
            col,
            "to pass",
            HEADING | BORDER_LEFT | BORDER_RIGHT | BORDER_BOTTOM | TEXT_WRAP,
        )
        ref = self.write_formula(
            row + 2,
            col,
            f"={self.VAR_MARKS_ALL_SHEETS}*0.5",  # TODO turn 0.5 into a config value
            HEADING | BORDER_ALL,
        )
        self.define_name(self.VAR_MARKS_TO_PASS, ref)

    def write_student_score_row(self, row, col, student_marks, email):
        for c, sheet in enumerate(self.sheets, start=col):
            if self.points_per == "exercise":
                # Marking by exercise, we take the sheet scores from the other
                # worksheet. We assume emails are in column A and sheetnames in
                # row 1, and that everything fits in A1:ZZ1000.
                formula = (
                    f"=INDEX({quote_sheetname(self.SHEET_NAME_INDIVIDUAL)}!B2:ZZ1000,"
                    f'MATCH("{email}", {quote_sheetname(self.SHEET_NAME_INDIVIDUAL)}!A2:A1000, 0),'
                    f'MATCH("{sheet}", {quote_sheetname(self.SHEET_NAME_INDIVIDUAL)}!B1:ZZ1, 0))'
                )
                self.write_formula(row, c, formula)
            else:
                marks = student_marks.get(sheet, "")
                self.write(row, c, convert_to_float_if_possible(marks))
        return OpenpyxlRangeRef(row, col, row, col + len(self.sheets) - 1)

    def write_student_total_marks(self, row, col, ref_individual_marks):
        ref_individual_marks.row_absolute = False
        formula = (
            f"=SUMPRODUCT({ref_individual_marks},{self.VAR_SHEET_WAS_MARKED})"
        )
        return self.write_formula(row, col, formula, BORDER_ALL)

    def write_student_missing_marks(self, row, col, ref_total_marks):
        ref_total_marks.row_absolute = False
        formula = f"=MAX(0,{self.VAR_MARKS_TO_PASS} - {ref_total_marks})"
        return self.write_formula(row, col, formula, BORDER_ALL)

    def write_student_average_marks(self, row, col, ref_individual_marks):
        ref_individual_marks.row_absolute = False
        formula = (
            f"=AVERAGE(IF(ISNUMBER({ref_individual_marks}) * SIGN({self.VAR_MARKS_BY_SHEET}) * {self.VAR_SHEET_WAS_MARKED}, "
            f"{ref_individual_marks} / {self.VAR_MARKS_BY_SHEET}, "
            f'"Ignoring plagiarism, bonus sheets, and sheets not submitted for the average"))'
        )
        return self.write_array_formula(row, col, formula, PERCENT | BORDER_ALL)

    def write_student_required_average_marks(self, row, col, ref_missing_marks):
        ref_missing_marks.row_absolute = False
        formula = (
            f"=IF({self.VAR_MARKS_FUTURE_SHEETS} > 0,"
            f"{ref_missing_marks}/{self.VAR_MARKS_FUTURE_SHEETS}, 10*SIGN({ref_missing_marks}))"
        )
        return self.write_formula(row, col, formula, PERCENT | BORDER_ALL)

    def write_student_required_improvement(
        self, row, col, ref_average_marks, ref_required_average
    ):
        ref_average_marks.row_absolute = False
        ref_required_average.row_absolute = False
        formula = f"={ref_required_average} - {ref_average_marks}"
        return self.write_formula(row, col, formula, PERCENT | BORDER_ALL)

    def write_student_summary_row(
        self, row, col, first_name, last_name, ref_individual_marks
    ):
        ref_individual_marks.row_absolute = False
        self.write(
            row, col, first_name, BORDER_TOP | BORDER_BOTTOM | BORDER_LEFT
        )
        self.write(
            row, col + 1, last_name, BORDER_TOP | BORDER_BOTTOM | BORDER_RIGHT
        )
        ref_total_marks = self.write_student_total_marks(
            row, col + 2, ref_individual_marks
        )
        ref_missing_marks = self.write_student_missing_marks(
            row, col + 3, ref_total_marks
        )
        ref_average_marks = self.write_student_average_marks(
            row, col + 4, ref_individual_marks
        )
        ref_required_average = self.write_student_required_average_marks(
            row, col + 5, ref_missing_marks
        )
        self.write_student_required_improvement(
            row, col + 6, ref_average_marks, ref_required_average
        )
        return OpenpyxlRangeRef(row, col, row, col + 7)

    def write_student_marks_table(self, row, col):
        # Headers
        headers = [
            "First Name",
            "Last Name",
            "Total Marks",
            "Missing Marks",
            "Avg Marks",
            "Required Avg",
            "Improve",
        ]
        self.write_row(row, col, headers, BOLD | BORDER_TOP | BORDER_BOTTOM)
        self.write_sheet_name_row(row, col + len(headers))
        # Content
        sorted_marks = sorted(
            self.students_marks.items(), key=sort_items_by_score
        )
        for r, (email, marks) in enumerate(sorted_marks, start=row + 1):
            first_name, last_name = self.email_to_name.get(
                email, ("Unknown", "Unknown")
            )
            ref_individual_marks = self.write_student_score_row(
                r, col + len(headers), marks, email
            )
            self.write_student_summary_row(
                r, col, first_name, last_name, ref_individual_marks
            )

        marks_min_row = row + 1
        marks_min_col = col + len(headers)
        marks_max_row = marks_min_row + len(sorted_marks) - 1
        marks_max_col = marks_min_col + len(sorted_marks) - 1
        return OpenpyxlRangeRef(
            marks_min_row, marks_min_col, marks_max_row, marks_max_col
        )

    def write_average_marks_by_sheet(self, row, col, ref_marks):
        self.merge_range(
            row, col, row, col + 1, "Average achieved", BOLD | RIGHT_ALIGNED
        )
        for c in range(ref_marks.min_column, ref_marks.max_column + 1):
            ref = OpenpyxlRangeRef(ref_marks.min_row, c, ref_marks.max_row, c)
            formula = f'=IFERROR(AVERAGE({ref}),"")'
            self.write_formula(row, c, formula, TWO_DIGIT_FLOAT)

    def write_color_key(self, row, col):
        self.write(row, col, "Color Key", BOLD)
        self.merge_range(row + 1, col, row + 1, col + 1, "Passed", GREEN)
        self.merge_range(row + 2, col, row + 2, col + 1, "Failed", RED)
        self.merge_range(
            row + 3, col, row + 3, col + 1, "2x Plagiarism", PLAGIARISM_RED
        )

        self.merge_range(
            row + 1,
            col + 4,
            row + 1,
            col + 5,
            "Does not need to improve\naverage if percentage is\nlower than:",
            GREEN | TEXT_WRAP,
        )
        self.merge_range(
            row + 2,
            col + 4,
            row + 2,
            col + 5,
            "Should improve average if\npercentage is between:",
            YELLOW | TEXT_WRAP,
        )
        self.merge_range(
            row + 3,
            col + 4,
            row + 3,
            col + 5,
            "Has to improve average by\nat least the following\npercentage:",
            RED | TEXT_WRAP,
        )

        cellref_min_good_improvement = self.write(
            row + 1, col + 6, -0.05, GREEN | PERCENT
        )
        self.define_name(
            self.VAR_MIN_GOOD_IMPROVEMENT, cellref_min_good_improvement
        )
        cellref_min_bad_improvement = self.write(
            row + 3, col + 6, 0.1, RED | PERCENT
        )
        self.define_name(
            self.VAR_MIN_BAD_IMPROVEMENT, cellref_min_bad_improvement
        )
        self.write(
            row + 2,
            col + 6,
            f'=TEXT({self.VAR_MIN_GOOD_IMPROVEMENT},"0%")& " to " & TEXT({self.VAR_MIN_BAD_IMPROVEMENT},"0%")',
            YELLOW,
        )

    def add_conditional_formatting_for_pass_fail(
        self, range, ref_individual_marks, ref_total_marks
    ):
        ref_individual_marks.row_absolute = False
        ref_individual_marks.column_absolute = True
        ref_total_marks.row_absolute = False
        ref_total_marks.column_absolute = True
        plagiarism_fail = (
            f'=COUNTIF({ref_individual_marks},"{strings.PLAGIARISM}") >= 2'
        )
        self.add_conditional_format(range, plagiarism_fail, PLAGIARISM_RED)
        impossible_pass = f"={ref_total_marks} + {self.VAR_MARKS_FUTURE_SHEETS} < {self.VAR_MARKS_TO_PASS}"
        self.add_conditional_format(range, impossible_pass, RED)
        already_passed = f"={ref_total_marks} >= {self.VAR_MARKS_TO_PASS}"
        self.add_conditional_format(range, already_passed, GREEN)

    def add_conditional_formatting_for_warnings(self, range, ref_improvement):
        ref_improvement.row_absolute = False
        ref_improvement.column_absolute = True
        self.add_conditional_format(
            range,
            f"={ref_improvement} <= {self.VAR_MIN_GOOD_IMPROVEMENT}",
            GREEN,
        )
        self.add_conditional_format(
            range,
            f"={ref_improvement} <= {self.VAR_MIN_BAD_IMPROVEMENT}",
            YELLOW,
        )
        self.add_conditional_format(
            range, f"={ref_improvement} > {self.VAR_MIN_BAD_IMPROVEMENT}", RED
        )

    def add_conditional_formatting_for_zebra_stripes(self, range):
        self.add_conditional_format(range, f"=ISEVEN(ROW())", GRAY)

    def write_summary_sheet(self):
        self.write_sheet_name_row(1, 8)
        self.write_available_marks_by_sheet(2, 6)
        self.write_was_marked_row(3, 6)
        self.write_available_marks_table(1, 2)
        self.write_required_marks_table(1, 1)
        ref_marks = self.write_student_marks_table(5, 1)
        self.write_average_marks_by_sheet(4, 6, ref_marks)
        self.write_color_key(ref_marks.max_row + 3, 1)

        # TODO: the following code makes a lot of assumptions about the position
        # of certain columns and is thus hard to untangle. Figure out a better
        # way to pass along this information.

        # Conditional Formatting
        def get_representative_student_ref(min_col, max_col=None):
            row = ref_marks.min_row
            return OpenpyxlRangeRef(row, min_col, row, max_col)

        def get_student_columns_ref(min_col, max_col=None):
            return OpenpyxlRangeRef(
                ref_marks.min_row, min_col, ref_marks.max_row, max_col
            )

        ref_student_name_columns = get_student_columns_ref(1, 2)
        ref_rep_individual_marks = get_representative_student_ref(
            ref_marks.min_column, ref_marks.max_column
        )
        ref_rep_total_marks = get_representative_student_ref(3)
        self.add_conditional_formatting_for_pass_fail(
            ref_student_name_columns,
            ref_rep_individual_marks,
            ref_rep_total_marks,
        )

        ref_student_average_columns = get_student_columns_ref(5, 7)
        ref_rep_improvement = get_representative_student_ref(7)
        self.add_conditional_formatting_for_warnings(
            ref_student_average_columns, ref_rep_improvement
        )

        ref_full_table = get_student_columns_ref(1, ref_marks.max_column)
        self.add_conditional_formatting_for_zebra_stripes(ref_full_table)

    def write_per_exercise_marks_sheet(self):
        sorted_marks = sorted(
            self.students_marks.items(), key=sort_items_by_score
        )
        tasks_on_sheet = defaultdict(set)
        for _, marks in sorted_marks:
            for sheet in self.sheets:
                exercise_marks = marks.get(sheet, {})
                for task in exercise_marks:
                    tasks_on_sheet[sheet].add(task)

        sheet_column = {}
        task_column = {}
        free_column = itertools.count(start=4)
        spacer_columns = [next(free_column)]
        for sheet in self.sheets:
            sheet_column[sheet] = next(free_column)
            for task in sorted(tasks_on_sheet[sheet]):
                task_column[sheet, task] = next(free_column)
            # Leave one column empty as a spacer
            spacer_columns.append(next(free_column))
        last_column = spacer_columns[-1]

        self.write(1, 1, "Email", BOLD | CENTERED | BORDER_BOTTOM)
        self.write(1, 2, "First Name", BOLD | CENTERED | BORDER_BOTTOM)
        self.write(1, 3, "Last Name", BOLD | CENTERED | BORDER_BOTTOM)
        for sheet in self.sheets:
            self.write(
                1, sheet_column[sheet], sheet, BOLD | CENTERED | BORDER_BOTTOM
            )
            for task in tasks_on_sheet[sheet]:
                self.write(
                    1,
                    task_column[sheet, task],
                    task,
                    BOLD | CENTERED | BORDER_BOTTOM,
                )
        for c in spacer_columns:
            self.write(1, c, "", GRAY | BORDER_BOTTOM | BORDER_RIGHT)
            for r in range(2, len(sorted_marks) + 2):
                self.write(r, c, "", GRAY | BORDER_RIGHT)

        for r, (email, marks) in enumerate(sorted_marks, start=2):
            first_name, last_name = self.email_to_name.get(
                email, ("Unknown", "Unknown")
            )
            self.write(r, 1, email)
            self.write(r, 2, first_name)
            self.write(r, 3, last_name)
            for sheet in self.sheets:
                sheet_marks = marks.get(sheet, {})
                for task, task_marks in sheet_marks.items():
                    self.write(
                        r,
                        task_column[sheet, task],
                        convert_to_float_if_possible(task_marks),
                    )

                col_sheet = sheet_column[sheet]
                min_col = col_sheet + 1
                max_col = min_col + max(1, len(tasks_on_sheet[sheet])) - 1
                ref_marks = OpenpyxlRangeRef(
                    r, min_col, r, max_col, row_absolute=False
                )
                formula = (
                    f'=IF(COUNTIF({ref_marks},"{strings.PLAGIARISM}") > 0,"{strings.PLAGIARISM}",'
                    + f'IF(COUNT({ref_marks})=0,"",SUM({ref_marks}))'
                )
                self.write_formula(r, col_sheet, formula, BORDER_RIGHT)
        ref = OpenpyxlRangeRef(1, 1, r, last_column)
        self.add_conditional_formatting_for_zebra_stripes(ref)

    def autofit_columns(self, min_width=5, max_width=50):
        def cell_as_str(cell):
            if cell.value is None:
                return ""
            elif (cell.row, cell.column) in merged_cells:
                return ""
            elif isinstance(cell.value, ArrayFormula) or str(
                cell.value
            ).startswith("="):
                return "000.00"
            else:
                return str(cell.value)

        merged_cells = set()
        for merged_range in self.worksheet.merged_cells.ranges:
            merged_cells.update(merged_range.cells)
        for col in self.worksheet.columns:
            max_length = max(len(cell_as_str(cell)) for cell in col)
            adjusted_width = max(min_width, min(max_width, max_length + 2))
            column_letter = get_column_letter(col[0].column)
            self.worksheet.column_dimensions[column_letter].width = (
                adjusted_width
            )

    def add_summary_sheet(self):
        self.worksheet = self.workbook.create_sheet(self.SHEET_NAME_SUMMARY, 0)
        self.write_summary_sheet()
        self.autofit_columns()

    def add_marks_per_exercise_sheet(self):
        self.worksheet = self.workbook.create_sheet(self.SHEET_NAME_INDIVIDUAL)
        self.write_per_exercise_marks_sheet()
        self.autofit_columns()


def load_marks_files(marks_dir: Path, _the_config: config.Config):
    """
    Loads the data of the individual marks files in the specified directory.
    """
    marks_files = marks_dir.glob(
        f"{strings.MARKS_FILE_PREFIX}"
        "*"
        f"{strings.INDIVIDUAL_MARKS_FILE_POSTFIX}.json"
    )
    if _the_config.points_per == "exercise":
        students_marks = defaultdict(lambda: defaultdict(dict))
        graded_sheet_names = defaultdict(list)
    else:
        students_marks = defaultdict(dict)
        graded_sheet_names = set()
    tutors = defaultdict(set)
    for file in marks_files:
        data = utils.read_json(file)
        sheet_name = data["adam_sheet_name"]
        marks = data["marks"]
        tutors[sheet_name].add(data["tutor_name"])
        if _the_config.points_per == "exercise":
            for email, exercises in marks.items():
                for exercise, mark in exercises.items():
                    if students_marks[email][sheet_name].get(exercise):
                        logging.warning(
                            f"{exercise} of sheet {sheet_name} is marked "
                            f"multiple times for {email}!"
                        )
                    students_marks[email][sheet_name][exercise] = mark
                    if exercise not in graded_sheet_names[sheet_name]:
                        graded_sheet_names[sheet_name].append(exercise)
        else:
            graded_sheet_names.add(sheet_name)
            for email, mark in marks.items():
                if students_marks[email].get(sheet_name):
                    logging.warning(
                        f"Sheet {sheet_name} is marked multiple times for "
                        f"{email}!"
                    )
                students_marks[email][sheet_name] = mark
    for sheet_name, tutor_list in tutors.items():
        for tutor in (
            _the_config.classes
            if _the_config.marking_mode == "static"
            else _the_config.tutor_list
        ):
            if tutor not in tutor_list:
                logging.warning(
                    f"There is no file from tutor {tutor} for sheet "
                    f"{sheet_name}!"
                )
    return students_marks, graded_sheet_names


def create_marks_summary_excel_file(
    _the_config: config.Config, marks_dir: Path
) -> None:
    """
    Generates an Excel file that summarizes the students' marks. Uses a path to
    a directory containing the individual marks files.
    """
    email_to_name = create_email_to_name_dict(_the_config.teams)
    students_marks, graded_sheet_names = load_marks_files(
        marks_dir, _the_config
    )

    workbook = Workbook()
    # Openpyxl creates an empty sheet in the workbook that we don't use and will
    # delete later.
    dummy_sheet = workbook.active

    builder = PointsSummarySheetBuilder(
        workbook,
        _the_config.points_per,
        _the_config.max_points_per_sheet,
        email_to_name,
        students_marks,
        graded_sheet_names,
    )
    if _the_config.points_per == "exercise":
        builder.add_marks_per_exercise_sheet()
    builder.add_summary_sheet()

    workbook.remove(dummy_sheet)
    workbook.save("points_summary_report.xlsx")


def summarize(_the_config: config.Config, args) -> None:
    """
    Generate an Excel file summarizing students' marks after the individual
    marks files have been collected in a directory.
    """
    if not args.marks_dir.is_dir():
        logging.critical("The given individual marks directory is not valid!")
    create_marks_summary_excel_file(_the_config, args.marks_dir)
