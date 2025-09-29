import json
import os
import pathlib
import shutil
import tempfile
import textwrap
from collections import defaultdict
from typing import Union
from zipfile import ZipFile

import openpyxl

from .. import config, errors, sheets, strings, submissions, utils
from ..teams import *


def extract_adam_zip(args) -> tuple[pathlib.Path, str]:
    """
    Unzips the given ADAM zip file to a directory named after the exercise sheet
    name provided by ADAM, or has the name given in the `target` argument.
    """
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract ADAM zip contents to a temporary directory (or move in case it
        # was extracted automatically, this happened on an Apple system).
        if args.adam_zip_path.is_file():
            # Unzip to the directory within the zip file.
            # Should be the name of the exercise sheet,
            # for example "Exercise Sheet 2".
            with ZipFile(args.adam_zip_path, mode="r") as zip_file:
                utils.filtered_extract(zip_file, pathlib.Path(temp_dir))
        else:
            # Assume the directory is an extracted ADAM zip.
            unzipped_path = pathlib.Path(args.adam_zip_path)
            unzipped_destination_path = (
                pathlib.Path(temp_dir) / unzipped_path.name
            )
            shutil.copytree(unzipped_path, unzipped_destination_path)

        # Check if the zip has the expected structure.
        children = list(pathlib.Path(temp_dir).iterdir())
        if len(children) != 1 or not children[0].is_dir():
            # We expect the zip to contain a single subdirectory with the
            # exercise sheet name given on ADAM.
            errors.unexpected_zip_structure(args.adam_zip_path)
        temp_sheet_root_dir = children[0]
        adam_sheet_name = temp_sheet_root_dir.name
        grand_children = list(pathlib.Path(temp_sheet_root_dir).iterdir())
        if (
            len(grand_children) != 2
            or not any(
                grand_child.is_file() and grand_child.suffix == ".xlsx"
                for grand_child in grand_children
            )
            or not any(grand_child.is_dir() for grand_child in grand_children)
        ):
            # Within its single subdirectory, we expect the zip to contain a
            # single subsubdirectory named either "Abgaben" or "Submissions" and
            # a single spreadsheet with information about the submissions.
            errors.unexpected_zip_structure(args.adam_zip_path)
        # Move extracted directory from the temporary directory to its final
        # location.
        destination = pathlib.Path(
            args.target if args.target else adam_sheet_name
        )
        if destination.exists():
            logging.critical(
                f"Extraction failed because the path '{destination}' exists"
                " already!"
            )
        sheet_root_dir = shutil.move(temp_sheet_root_dir, destination)
    # Flatten intermediate directory.
    sub_directories = [
        sub_directory
        for sub_directory in sheet_root_dir.iterdir()
        if sub_directory.is_dir()
    ]
    assert len(sub_directories) == 1
    utils.move_content_and_delete(sub_directories[0], sheet_root_dir)
    return sheet_root_dir, adam_sheet_name


def mark_irrelevant_team_dirs(
    _the_config: config.Config, sheet: sheets.Sheet
) -> None:
    """
    Indicate which team directories do not have to be marked by adding the
    `DO_NOT_MARK_PREFIX` to their directory name.
    """
    for submission in sheet.get_all_team_submission_info():
        if not submission.relevant:
            shutil.move(
                submission.root_dir,
                submission.root_dir.with_name(
                    strings.DO_NOT_MARK_PREFIX + submission.root_dir.name
                ),
            )


def rename_team_dirs(sheet: sheets.Sheet) -> None:
    """
    The team directories are renamed to: team_id_LastName1-LastName2
    The team ID can be helpful to identify a team on the ADAM web interface.
    """
    for submission in sheet.get_all_team_submission_info():
        team_key = submission.team.get_team_key()
        team_dir = pathlib.Path(
            shutil.move(
                submission.root_dir, submission.root_dir.with_name(team_key)
            )
        )


def flatten_team_dirs(sheet: sheets.Sheet) -> None:
    """
    There can be multiple directories within a "Team 00000" directory. This
    happens when multiple members of the team upload solutions. Sometimes, only
    one directory contains submitted files, in this case we remove the empty
    ones silently. In case multiple submissions exist, we put the files within
    them next to each other and print a warning.
    """
    for submission in sheet.get_all_team_submission_info():
        # Remove empty subdirectories.
        for team_submission_dir in submission.root_dir.iterdir():
            if (
                team_submission_dir.is_dir()
                and len(list(team_submission_dir.iterdir())) == 0
            ):
                team_submission_dir.rmdir()
        # Store the list of team submission directories in variable, because the
        # generator may include subdirectories of team submission directories
        # that have already been flattened.
        team_submission_dirs = [
            path
            for path in submission.root_dir.iterdir()
            if not path.name == strings.SUBMISSION_INFO_FILE_NAME
        ]
        if len(team_submission_dirs) > 1:
            logging.warning(
                f"There are multiple submissions for group '{submission.root_dir.name}'!"
            )
        if len(team_submission_dirs) < 1:
            logging.warning(
                f"The submission of group '{submission.root_dir.name}' is empty!"
            )
        for team_submission_dir in team_submission_dirs:
            if team_submission_dir.is_dir():
                utils.move_content_and_delete(
                    team_submission_dir, submission.root_dir
                )


def unzip_internal_zips(sheet: sheets.Sheet) -> None:
    """
    If multiple files are uploaded to ADAM, the submission becomes a single zip
    file. Here we extract this zip. I'm not sure if nested zip files are also
    extracted. Additionally, we flatten the directory as long as a level only
    consists of a single directory.
    """
    for submission in sheet.get_all_team_submission_info():
        for zip_file in submission.root_dir.glob("**/*.zip"):
            with ZipFile(zip_file, mode="r") as zf:
                utils.filtered_extract(zf, zip_file.parent)
            os.remove(zip_file)
        sub_dirs = [
            path
            for path in submission.root_dir.iterdir()
            if not path.name == strings.SUBMISSION_INFO_FILE_NAME
        ]
        while len(sub_dirs) == 1 and sub_dirs[0].is_dir():
            utils.move_content_and_delete(sub_dirs[0], submission.root_dir)
            sub_dirs = [
                path
                for path in submission.root_dir.iterdir()
                if not path.name == strings.SUBMISSION_INFO_FILE_NAME
            ]


def create_marks_file(
    _the_config: config.Config, sheet: sheets.Sheet, args
) -> None:
    """
    Write a json file to add the marks for all relevant teams and exercises.
    """
    exercise_dict: Union[str, dict[str, str]] = ""
    if _the_config.points_per == "exercise":
        if _the_config.marking_mode == "static":
            exercise_dict = {
                f"exercise_{i}": "" for i in range(1, args.num_exercises + 1)
            }
        elif _the_config.marking_mode == "exercise":
            exercise_dict = {f"exercise_{i}": "" for i in sheet.exercises}
    else:
        exercise_dict = ""

    marks_dict = {}
    for submission in sorted(sheet.get_relevant_submissions()):
        team_key = submission.team.get_team_key()
        marks_dict.update({team_key: exercise_dict})

    with open(
        sheet.get_marks_file_path(_the_config), "w", encoding="utf-8"
    ) as marks_json:
        json.dump(marks_dict, marks_json, indent=4, ensure_ascii=False)


def create_feedback_directories(
    _the_config: config.Config, sheet: sheets.Sheet, plain: bool
) -> None:
    """
    Create a directory for every team that should be corrected by the tutor
    specified in the config. A copy of every file is prefixed and placed
    in the feedback folder. If there are multiple PDFs, we keep the file
    names as submitted. The idea is that feedback can be added to these
    copies directly and files without feedback can simply be deleted.
    """
    for submission in sheet.get_relevant_submissions():
        feedback_dir = submission.get_feedback_dir()
        feedback_dir.mkdir()

        feedback_file_name = sheet.get_feedback_file_name(_the_config)
        if not _the_config.xopp:
            feedback_pdf_name = feedback_file_name + ".pdf"
            pdf_files = list(submission.root_dir.glob("*.pdf"))
            if len(pdf_files) == 1:
                shutil.copy(pdf_files[0], feedback_dir / feedback_pdf_name)
            elif len(pdf_files) > 1:
                logging.warning(
                    f"There are multiple PDFs in the "
                    f"submission directory {submission.root_dir}."
                )
                for pdf in pdf_files:
                    shutil.copy(pdf, feedback_dir)

        # Copy non-pdf submission files into feedback directory with added
        # prefix.
        if not plain:
            for submission_file in submission.root_dir.glob("*"):
                if (
                    submission_file.is_dir()
                    or submission_file.suffix == ".pdf"
                    or submission_file.name == strings.SUBMISSION_INFO_FILE_NAME
                ):
                    continue
                this_feedback_file_name = (
                    feedback_file_name + "_" + submission_file.name
                )
                shutil.copy(
                    submission_file, feedback_dir / this_feedback_file_name
                )


def generate_xopp_files(
    sheet: sheets.Sheet, _the_config: config.Config
) -> None:
    """
    Generate xopp files in the feedback directories that point to the pdfs
    in the submission directory.
    """
    from pypdf import PdfReader

    def write_to_file(f, string):
        f.write(textwrap.dedent(string))

    logging.info("Generating .xopp files...")
    for submission in sheet.get_relevant_submissions():
        feedback_dir = submission.get_feedback_dir()
        pdf_paths = list(submission.root_dir.glob("*.pdf"))
        if len(pdf_paths) > 1:
            logging.warning(
                "There are multiple PDFs in the submission directory "
                f"{submission.root_dir}."
            )
        for pdf_path in pdf_paths:
            file_name = pdf_path.name
            if len(pdf_paths) == 1:
                file_name = sheet.get_feedback_file_name(_the_config) + ".pdf"
            pages = PdfReader(pdf_path).pages
            xopp_path = (feedback_dir / file_name).with_suffix(".xopp")
            if xopp_path.is_file():
                logging.warning(
                    "Skipping .xopp file generation for "
                    f"{submission.root_dir.name}: xopp file exists."
                )
                continue
            xopp_file = open(xopp_path, "w", encoding="utf-8")
            for i, page in enumerate(pages, start=1):
                width = page.mediabox.width
                height = page.mediabox.height
                if i == 1:  # Special entry for first page
                    write_to_file(
                        xopp_file,
                        f"""\
                        <?xml version="1.0" standalone="no"?>
                        <xournal creator="Xournal++ 1.1.1" fileversion="4">
                        <title>Xournal++ document - see https://github.com/xournalpp/xournalpp</title>
                        <page width="{width}" height="{height}">
                        <background type="pdf" domain="absolute" filename="{pdf_path.resolve()}" pageno="{i}"/>
                        <layer/>
                        </page>""",  # noqa
                    )
                else:
                    write_to_file(
                        xopp_file,
                        f"""\
                        <page width="{width}" height="{height}">
                        <background type="pdf" pageno="{i}"/>
                        <layer/>
                        </page>""",
                    )
            write_to_file(xopp_file, "</xournal>")
            xopp_file.close()
    logging.info("Done generating .xopp files.")


def print_missing_submissions(
    _the_config: config.Config, sheet: sheets.Sheet
) -> None:
    """
    Print all teams that are listed in the config file, but whose submission is
    not present in the zip downloaded from ADAM.
    """
    teams_who_submitted = [
        submission.team for submission in sheet.get_all_team_submission_info()
    ]
    students_who_submitted = [
        member for team in teams_who_submitted for member in team.members
    ]
    # Also checks if the team has been restructured
    missing_teams = [
        team
        for team in _the_config.teams
        if team not in teams_who_submitted
        and not any(member in students_who_submitted for member in team.members)
    ]
    if missing_teams:
        logging.info("There are no submissions for the following team(s):")
        for missing_team in missing_teams:
            print(f"* {missing_team}")


def set_relevance_for_submission_teams(
    _the_config: config.Config, submission_teams: dict[str, Team]
) -> dict[str, bool]:
    """
    Determines the value of 'relevant' of the submission.json files for the
    submission teams. Returns a dictionary with the team IDs as keys and a
    boolean as value indicating if the team is relevant or not.
    """
    team_relevance_dict = {}
    student_email_to_tutor = create_student_email_to_tutor_dict(_the_config)
    team_to_tutors = create_submission_team_to_tutors_dict(
        list(submission_teams.values()), student_email_to_tutor, _the_config
    )
    for team_id, tutors in team_to_tutors.items():
        if len(tutors) != 1:
            if _the_config.tutor_name in tutors:
                team_relevance_dict[team_id] = True
                if (
                    _the_config.marking_mode == "static"
                    and len(_the_config.classes.keys()) > 1
                ):
                    logging.warning(
                        f"Team {submission_teams[team_id]} is now assigned to "
                        f"tutors {tutors}.\n"
                        "Please contact the other tutors to decide who will "
                        "mark this team. Update the shared config file and "
                        "share it with your fellow tutors as well as with the "
                        "teaching assistant.\n"
                        "If you will not mark this team, then:\n"
                        "* Set the value of relevant to false in the "
                        "submission.json file of the team directory.\n"
                        "* Remove the team from the points_*.json file."
                    )
            else:
                team_relevance_dict[team_id] = False
        else:
            if _the_config.tutor_name in tutors:
                team_relevance_dict[team_id] = True
            else:
                team_relevance_dict[team_id] = False
    return team_relevance_dict


def create_submission_team_to_tutors_dict(
    submission_teams: list[Team],
    student_email_to_tutor: dict[str, set[str]],
    _the_config: config.Config,
) -> dict[str, set[str]]:
    """
    Create a dictionary that maps submission team IDs to a set of assigned
    tutors.
    """
    team_to_tutors = defaultdict(set)
    for team in submission_teams:
        if is_new_team(_the_config, team):
            team_to_tutors[team.adam_id] = (
                set(_the_config.classes.keys())
                if (_the_config.marking_mode == "static")
                else (set(_the_config.tutor_list))
            )
        else:
            for member in team.members:
                if member.email in student_email_to_tutor:
                    for tutor in student_email_to_tutor[member.email]:
                        team_to_tutors[team.adam_id].add(tutor)
    return team_to_tutors


def create_student_email_to_tutor_dict(
    _the_config: config.Config,
) -> dict[str, set[str]]:
    """
    Creates a dictionary that maps email addresses of students in the config
    to a set of assigned tutors.
    """
    email_to_tutor_dict = defaultdict(set)
    if _the_config.marking_mode == "static":
        for tutor, teams in _the_config.classes.items():
            for team in teams:
                for member in team.members:
                    email_to_tutor_dict[member.email].add(tutor)
    elif _the_config.marking_mode == "exercise":
        for team in _the_config.teams:
            for member in team.members:
                for tutor in _the_config.tutor_list:
                    email_to_tutor_dict[member.email].add(tutor)
    else:
        errors.unsupported_marking_mode_error(_the_config.marking_mode)
    return email_to_tutor_dict


def is_new_team(_the_config: config.Config, submission_team: Team) -> bool:
    """
    Checks if a given submission team is a new team consisting only of new
    students.
    """
    students_in_config_teams = [
        member for team in _the_config.teams for member in team.members
    ]
    return all(
        member not in students_in_config_teams
        for member in submission_team.members
    )


def is_restructured_submission_team(
    _the_config: config.Config, submission_team: Team
) -> bool:
    """
    Checks if the given submission team is structured differently in the config.
    This ignores new submission teams consisting only of new students.
    """
    students_in_config_teams = [
        member for team in _the_config.teams for member in team.members
    ]
    return submission_team not in _the_config.teams and any(
        member in students_in_config_teams for member in submission_team.members
    )


def get_original_config_teams(
    _the_config: config.Config, submission_team: Team
) -> list[Team]:
    """
    Finds all the config teams that contain a member of the given
    submission team.
    """
    original_config_teams = []
    for member in submission_team.members:
        for config_team in _the_config.teams:
            if (
                member in config_team.members
                and config_team not in original_config_teams
            ):
                original_config_teams.append(config_team)
    return original_config_teams


def is_in_config_teams(_the_config: config.Config, student: Student) -> bool:
    """
    Checks if a student appears in the config teams.
    """
    return student in [
        member for team in _the_config.teams for member in team.members
    ]


def validate_team_size(
    max_team_size: int, submission_teams: list[Team]
) -> None:
    """
    Checks if the team size of the submission teams does not exceed the
    maximum allowed team size.
    """
    teams = [
        team for team in submission_teams if len(team.members) > max_team_size
    ]
    if teams:
        logging.warning(
            "There are submission teams that have " "more members than allowed:"
        )
    for team in teams:
        print(f"* {team}")


def validate_teams(
    _the_config: config.Config, submission_teams: list[Team]
) -> None:
    """
    Checks if submission teams are organized differently in the config
    and if there are new teams consisting only of new students.
    """
    new_submission_teams = [
        submission_team
        for submission_team in submission_teams
        if is_restructured_submission_team(_the_config, submission_team)
    ]
    if new_submission_teams:
        logging.warning(
            "There are submission teams that are structured differently in the "
            "config."
        )
        print(strings.SEPARATOR_LINE)
        for new_submission_team in new_submission_teams:
            print("New submission team:")
            print(f"* {new_submission_team}")
            original_teams = get_original_config_teams(
                _the_config, new_submission_team
            )
            if original_teams:
                print("Related config teams:")
                for original_team in original_teams:
                    print(f"* {original_team}")
            new_students = [
                member
                for member in new_submission_team.members
                if not is_in_config_teams(_the_config, member)
            ]
            if new_students:
                print(
                    "Members of the new submission team that do not appear in "
                    "the config:"
                )
                for student in new_students:
                    print(f"* {student}")
            print(strings.SEPARATOR_LINE)
    new_teams = [
        submission_team
        for submission_team in submission_teams
        if is_new_team(_the_config, submission_team)
    ]
    if new_teams:
        logging.warning(
            "There are completely new teams where all members "
            "are not listed in the config:"
        )
        for new_team in new_teams:
            print(f"* {new_team}")


def create_all_submission_info_files(
    _the_config: config.Config,
    submission_teams: dict[str, Team],
    team_relevance_dict: dict[str, bool],
    sheet_root_dir: pathlib.Path,
) -> None:
    """
    Creates the submission info JSON files in all team directories.
    """
    for team_dir in sheet_root_dir.iterdir():
        if team_dir.is_dir():
            team_id = team_dir.name.split(" ")[1]
            team = submission_teams[team_id]
            submissions.create_submission_info_file(
                _the_config, team, team_relevance_dict[team_id], team_dir
            )


def use_names_from_config(
    config_teams: list[Team], submission_teams: dict[str, Team]
) -> None:
    """
    Changes the names of the students in the submission teams to the names
    defined in the config, if available.
    """
    email_to_name_dict = create_email_to_name_dict(config_teams)
    for team in submission_teams.values():
        for member in team.members:
            if member.email in email_to_name_dict:
                member.first_name, member.last_name = email_to_name_dict[
                    member.email
                ]


def read_teams_from_adam_spreadsheet(
    sheet_root_dir: pathlib.Path,
) -> dict[str, Team]:
    """
    Reads the teams from the ADAM Excel spreadsheet and returns a dictionary
    with the team IDs as keys and the teams as values.
    """
    excel_files = list(sheet_root_dir.glob("*.xlsx"))
    if not excel_files:
        logging.critical("No ADAM Excel spreadsheet found.")
    wb = openpyxl.load_workbook(excel_files[0])
    sheet = wb.active
    col_last_name = 0
    col_first_name = 1
    col_email = 2
    col_team_id = 4
    teams_data = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        team_id = str(row[col_team_id])
        first_name = row[col_first_name]
        last_name = row[col_last_name]
        email = row[col_email]
        teams_data[team_id].append((first_name, last_name, email))
    for team in teams_data.values():
        team.sort()
    teams = {}
    for team_id, team in teams_data.items():
        teams[team_id] = Team([Student(*student) for student in team], team_id)
    return teams


def init(_the_config: config.Config, args) -> None:
    """
    Prepares the directory structure holding the submissions.
    """
    # Catch wrong combinations of marking_mode/points_per/-n/-e.
    # Not possible earlier because marking_mode and points_per are given by the
    # config file.
    if _the_config.points_per == "exercise":
        if _the_config.marking_mode == "exercise" and not args.exercises:
            logging.critical(
                "You must provide a list of exercise numbers to be marked with "
                "the '-e' flag, for example '-e 1 3 4'. In case the '-e' flag "
                "is the last option before the ADAM zip path, make sure to "
                "separate exercise numbers from the path using '--'."
            )
        if _the_config.marking_mode == "static" and not args.num_exercises:
            logging.critical(
                "You must provide the number of exercises in the sheet with "
                "the '-n' flag, for example '-n 5'."
            )
    else:  # points per sheet
        if _the_config.marking_mode == "exercise":
            logging.critical(
                "Points must be given per exercise if marking is done per "
                "exercise. Set the value of 'poins_per' to 'exercise' or "
                "change the marking mode."
            )
        if args.num_exercises or args.exercises:
            logging.warning(
                "'points_per' is 'sheet', so the flags '-n' and '-e' are "
                "ignored."
            )

    # Sort the list exercises of flag "-e" to make printing nicer later.
    if args.exercises:
        args.exercises.sort()

    # The adam zip file is expected to have the following structure:
    # <ADAM Exercise Sheet Name>.zip
    # └── <ADAM Exercise Sheet Name>
    #     └── Abgaben
    #         ├── Team 12345
    #         .   └── Muster_Hans_hans.muster@unibas.ch_000000
    #         .       └── submission.pdf or submission.zip
    sheet_root_dir, adam_sheet_name = extract_adam_zip(args)

    # Structure at this point:
    # <sheet_root_dir>
    # ├── Team 12345
    # .   └── Muster_Hans_hans.muster@unibas.ch_000000
    # .       └── submission.pdf or submission.zip
    submission_teams = read_teams_from_adam_spreadsheet(sheet_root_dir)
    use_names_from_config(_the_config.teams, submission_teams)
    validate_team_size(
        _the_config.max_team_size, list(submission_teams.values())
    )
    if _the_config.marking_mode == "static":
        validate_teams(_the_config, list(submission_teams.values()))
    team_relevance_dict = set_relevance_for_submission_teams(
        _the_config, submission_teams
    )
    create_all_submission_info_files(
        _the_config, submission_teams, team_relevance_dict, sheet_root_dir
    )
    sheet = sheets.create_sheet_info_file(
        sheet_root_dir, adam_sheet_name, _the_config, args.exercises
    )
    print_missing_submissions(_the_config, sheet)

    # Structure at this point:
    # <sheet_root_dir>
    # ├── Team 12345
    # .   ├── Muster_Hans_hans.muster@unibas.ch_000000
    # .   │   └── submission.pdf or submission.zip
    # .   └── submission.json
    # └── sheet.json
    rename_team_dirs(sheet)

    # Structure at this point:
    # <sheet_root_dir>
    # ├── 12345_Muster-Meier-Mueller
    # .   ├── Muster_Hans_hans.muster@unibas.ch_000000
    # .   │   └── submission.pdf or submission.zip
    # .   └── submission.json
    # └── sheet.json
    flatten_team_dirs(sheet)

    # Structure at this point:
    # <sheet_root_dir>
    # ├── 12345_Muster-Meier-Mueller
    # .   ├── submission.pdf or submission.zip
    # .   └── submission.json
    # └── sheet.json
    unzip_internal_zips(sheet)

    # From here on, we need information about relevant teams.
    mark_irrelevant_team_dirs(_the_config, sheet)

    if _the_config.use_marks_file:
        create_marks_file(_the_config, sheet, args)

    create_feedback_directories(_the_config, sheet, args.plain)

    # Structure at this point:
    # <sheet_root_dir>
    # ├── 12345_Muster-Meier-Mueller
    # .   ├── feedback
    # .   │   ├── feedback.pdf
    # .   │   └── feedback_copy-of-submitted-file.cc
    # .   ├── submission.pdf or submission files
    # .   └── submission.json
    # ├── sheet.json
    # └── points.json
    if _the_config.xopp:
        generate_xopp_files(sheet, _the_config)
